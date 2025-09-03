import openpyxl
import re
import pandas as pd
from copy import copy

class LabParser:
    def __init__(self, input_file, reference_file, output_file):
        self.input_file = input_file
        self.reference_file = reference_file
        self.output_file = output_file
        self.df = pd.read_excel(input_file, engine='openpyxl')
        self.df_exploded = None

    @staticmethod
    def split_by_result_anchor(text):
        if not isinstance(text, str):
            return []
        pattern = r'.*?отрицательный\)'
        blocks = re.findall(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        return [b.strip(' ;\n') for b in blocks if b.strip(' ;\n')]

    @staticmethod
    def parse_lab_block(block):
        reg = re.compile(
            r'^(?P<lab>.+?),\s*эксп\. № (?P<num>.+? от .+?г\.?)\s*\((?P<result>.+)\)$',
            re.DOTALL
        )
        m = reg.match(block)
        if m:
            return m.group('lab').strip(), m.group('num').strip(), m.group('result').strip()
        else:
            return '', '', block.strip()

    def explode_and_parse(self):
        self.df['lab_blocks'] = self.df['Результат лабораторного исследования'].apply(
            lambda x: self.split_by_result_anchor(x) if pd.notna(x) else []
        )
        df_exploded = self.df.explode('lab_blocks').reset_index(drop=True)
        df_exploded = df_exploded[df_exploded['lab_blocks'].notna() & (df_exploded['lab_blocks'] != '')]
        df_exploded[['Лаборатория', 'Номер и дата лабораторного исследования',
                     'Результат лабораторного исследования']] = (
            df_exploded['lab_blocks'].apply(lambda x: pd.Series(self.parse_lab_block(x)))
        )
        self.df_exploded = df_exploded

    @staticmethod
    def reorder_and_save_df(df, out_file):
        desired_order = [
            '№ п/п', 'Наименование продукции', 'Производитель',
            'Лаборатория', 'Номер и дата лабораторного исследования', 'Результат лабораторного исследования',
            'Номер ТТН', 'Дата ТТН'
        ]
        cols = [col for col in desired_order if col in df.columns]
        df[cols].to_excel(out_file, index=False)

    def reorder_and_save(self):
        LabParser.reorder_and_save_df(self.df_exploded, self.output_file)

    def apply_formatting(self):
        """Форматирует self.output_file по self.reference_file"""
        LabParser.apply_formatting_to_file(self.output_file, self.reference_file)

    @staticmethod
    def apply_formatting_to_file(out_file, reference_file="Формат_выгрузки.xlsx"):
        """
        Форматирует out_file по эталонному reference_file.
        """
        wb_ref = openpyxl.load_workbook(reference_file)
        ws_ref = wb_ref.active

        wb_out = openpyxl.load_workbook(out_file)
        ws_out = wb_out.active

        for col in ws_ref.column_dimensions:
            ws_out.column_dimensions[col].width = ws_ref.column_dimensions[col].width
        for row in ws_ref.row_dimensions:
            ws_out.row_dimensions[row].height = ws_ref.row_dimensions[row].height
        for i, ref_cell in enumerate(ws_ref[1], 1):
            out_cell = ws_out.cell(row=1, column=i)
            out_cell.font = copy(ref_cell.font)
            out_cell.fill = copy(ref_cell.fill)
            out_cell.border = copy(ref_cell.border)
            out_cell.alignment = copy(ref_cell.alignment)
            out_cell.number_format = copy(ref_cell.number_format)
        for row_idx in range(2, ws_out.max_row + 1):
            for col_idx, ref_cell in enumerate(ws_ref[2], 1):
                out_cell = ws_out.cell(row=row_idx, column=col_idx)
                out_cell.font = copy(ref_cell.font)
                out_cell.fill = copy(ref_cell.fill)
                out_cell.border = copy(ref_cell.border)
                out_cell.alignment = copy(ref_cell.alignment)
                out_cell.number_format = copy(ref_cell.number_format)
        ws_out.freeze_panes = 'A2'
        wb_out.save(out_file)

    def full_parse_and_format(self):
        self.explode_and_parse()
        self.reorder_and_save()
        self.apply_formatting()
